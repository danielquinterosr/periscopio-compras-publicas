# src/etl.py
import os
import re
import json
import time
import math
import glob
import hashlib
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook

from src.scoring import load_rules, total_score

import random
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# sesión global con retries
_SESSION = None

# -----------------------------
# Config general
# -----------------------------
MP_BASE = "https://api.mercadopublico.cl/servicios/v1/publico"
LICITACIONES_LIST_URL = f"{MP_BASE}/licitaciones.json"
# Nota: el detalle también se obtiene vía licitaciones.json, pero con ?codigo=... (no id)
# Fuente: documentación oficial del servicio público de licitaciones. :contentReference[oaicite:1]{index=1}

DOCS_DATA_DIR = "docs/data"
DATA_DIR = "data"

OUT_OPPS = os.path.join(DOCS_DATA_DIR, "opportunities.json")
OUT_META = os.path.join(DOCS_DATA_DIR, "meta.json")
OUT_REGISTRY = os.path.join(DOCS_DATA_DIR, "opportunities_registry.json")

DEFAULT_ONU_XLSX = os.getenv("ONU_RUBROS_XLSX", "config/Listado_rubros_ONU.xlsx")

# Filtrado inicial (MVP) por "Nivel1" ONU (ajustable)
DEFAULT_DENY_NIVEL1 = {
    "Equipamiento y suministros médicos",
    "Medicamentos y productos farmacéuticos",
    # opcional (descomenta si te está metiendo ultrasonidos/equipos)
    "Equipamiento para laboratorios",
}

# Compra Ágil: link directo a ficha
def compra_agil_url(code: str) -> str:
    return f"https://buscador.mercadopublico.cl/ficha?code={code}"

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

def safe_int(x: Any, default: int = 0) -> int:
    try:
        return int(float(x))
    except Exception:
        return default

def safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        return float(x)
    except Exception:
        return None

def ensure_dirs() -> None:
    os.makedirs(DOCS_DATA_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

# -----------------------------
# GitHub reviewed ids (issues label reviewed)
# -----------------------------
def fetch_reviewed_ids_from_github(repo: str, token: str) -> List[str]:
    """
    Busca issues con label 'reviewed' y extrae IDs desde:
      - título: "Reviewed: <id>"
      - body: línea "- ID: <id>"
    """
    if not repo or not token:
        return []

    reviewed_ids: List[str] = []
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "periscopio-bot",
    })

    page = 1
    while True:
        url = f"https://api.github.com/repos/{repo}/issues"
        params = {"state": "all", "labels": "reviewed", "per_page": 100, "page": page}
        r = s.get(url, params=params, timeout=30)
        if r.status_code != 200:
            break

        items = r.json() or []
        if not items:
            break

        for it in items:
            title = (it.get("title") or "").strip()
            body = (it.get("body") or "")

            m = re.match(r"^Reviewed:\s*(.+)$", title, flags=re.IGNORECASE)
            if m:
                reviewed_ids.append(m.group(1).strip())
                continue

            m2 = re.search(r"^-?\s*ID:\s*(.+)$", body, flags=re.IGNORECASE | re.MULTILINE)
            if m2:
                reviewed_ids.append(m2.group(1).strip())

        page += 1

    # únicos preservando orden
    seen = set()
    out = []
    for x in reviewed_ids:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out

# -----------------------------
# ONU Rubros mapping
# -----------------------------
def load_onu_mapping(path_xlsx: str) -> Dict[str, Dict[str, str]]:
    """
    Devuelve dict:
      IDCategoria(str) -> {"Nivel1":..., "Nivel2":..., "Nivel3":..., "Nivel4":..., "Categoria":...}
    Si el archivo no existe, retorna {}.
    """
    if not os.path.exists(path_xlsx):
        return {}

    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    required = ["IDCategoria", "Nivel1", "Nivel2", "Nivel3", "Nivel4", "Categoria"]
    if not all(r in idx for r in required):
        return {}

    mapping: Dict[str, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat = row[idx["IDCategoria"]]
        if cat is None:
            continue
        cat_s = str(cat).strip()
        mapping[cat_s] = {
            "Nivel1": str(row[idx["Nivel1"]] or "").strip(),
            "Nivel2": str(row[idx["Nivel2"]] or "").strip(),
            "Nivel3": str(row[idx["Nivel3"]] or "").strip(),
            "Nivel4": str(row[idx["Nivel4"]] or "").strip(),
            "Categoria": str(row[idx["Categoria"]] or "").strip(),
        }
    return mapping

# -----------------------------
# Licitaciones: listado + detalle
# -----------------------------
def _get_session():
    global _SESSION
    if _SESSION is not None:
        return _SESSION

    s = requests.Session()
    retries = Retry(
        total=6,                 # total reintentos
        connect=6,
        read=6,
        backoff_factor=1.2,      # 1.2s, 2.4s, 4.8s...
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retries, pool_connections=10, pool_maxsize=10)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    _SESSION = s
    return _SESSION


def mp_get(url: str, params: dict, timeout=(15, 90)) -> dict:
    """
    timeout = (connect_timeout, read_timeout)
    """
    s = _get_session()

    # jitter pequeño para no golpear siempre igual (reduce 429 y micro-bloqueos)
    time.sleep(0.15 + random.random() * 0.25)

    r = s.get(url, params=params, timeout=timeout)

    # Si el servidor devuelve HTML/otro, lo tratamos como error auditable
    ct = (r.headers.get("content-type") or "").lower()
    if r.status_code >= 400:
        raise RuntimeError(f"MP GET {url} -> {r.status_code}: {r.text[:200]}")

    if "application/json" not in ct:
        # MercadoPublico a veces devuelve cosas raras cuando está caído
        raise RuntimeError(f"MP GET {url} content-type inesperado={ct}. preview={r.text[:200]}")

    return r.json()

def fetch_licitaciones_list(ticket: str, page: int) -> Dict[str, Any]:
    # Este endpoint acepta varios parámetros; mantenemos solo los que ya venías usando.
    # Si más adelante quieres filtrar por estado, fecha, región, etc., lo agregamos.
    params = {"ticket": ticket, "pagina": page}
    return mp_get(LICITACIONES_LIST_URL, params=params)

def fetch_licitacion_detail(ticket: str, codigo: str) -> Optional[Dict[str, Any]]:
    """
    Detalle por código: ?codigo=<...>&ticket=<...>
    Importante: NO usar id=... (eso da 400 "Nombre de parametro no válido"). :contentReference[oaicite:2]{index=2}
    """
    params = {"ticket": ticket, "codigo": codigo}
    try:
        return mp_get(LICITACIONES_LIST_URL, params=params)
    except Exception:
        return None

def extract_category_codes_from_detail(detail_json: Dict[str, Any]) -> List[str]:
    """
    Intenta sacar:
      Listado[0].Items.Listado[*].CodigoCategoria
    y devuelve lista única (strings).
    """
    out: List[str] = []
    try:
        listado = (detail_json or {}).get("Listado") or []
        if not listado:
            return []
        first = listado[0] or {}
        items = (first.get("Items") or {}).get("Listado") or []
        for it in items:
            if not isinstance(it, dict):
                continue
            code = it.get("CodigoCategoria")
            if code is None:
                continue
            code_s = str(code).strip()
            if code_s:
                out.append(code_s)
    except Exception:
        return []

    seen = set()
    uniq = []
    for x in out:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq

# -----------------------------
# Compra Ágil: parse excel
# -----------------------------
def read_compra_agil_excel(path_xlsx: str) -> List[Dict[str, Any]]:
    """
    Espera headers (como ya validaste):
      ['ID','Nombre','Fecha de Publicación','Fecha de cierre','Organismo','Unidad','Monto Disponible','Moneda','Estado']
    """
    if not os.path.exists(path_xlsx):
        return []

    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    h = {name: i for i, name in enumerate(headers)}

    required = ["ID", "Nombre", "Fecha de Publicación", "Fecha de cierre", "Organismo", "Monto Disponible", "Estado"]
    if not all(k in h for k in required):
        return []

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = r[h["ID"]]
        title = r[h["Nombre"]]
        pub = r[h["Fecha de Publicación"]]
        close = r[h["Fecha de cierre"]]
        org = r[h["Organismo"]]
        amt = r[h["Monto Disponible"]]
        status = r[h["Estado"]]

        if not code or not title:
            continue

        def dt_to_iso(x: Any) -> Optional[str]:
            if x is None:
                return None
            if isinstance(x, datetime):
                # Excel suele traer naive; asumimos hora local ya “como viene”
                return x.replace(tzinfo=None).isoformat()
            # algunos excels vienen como string
            try:
                return str(x)
            except Exception:
                return None

        rows.append({
            "source": "compra_agil",
            "id": str(code).strip(),
            "title": str(title).strip(),
            "buyer": str(org).strip() if org else None,
            "status": str(status).strip() if status else None,
            "amount_clp": safe_float(amt),
            "published_at": dt_to_iso(pub),
            "close_at": dt_to_iso(close),
            "url": compra_agil_url(str(code).strip()),
        })
    return rows

# -----------------------------
# Registry (histórico + reviewed)
# -----------------------------
def load_json_file(path: str, default: Any) -> Any:
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json_file(path: str, obj: Any) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)

def upsert_registry(reg: Dict[str, Any], opp: Dict[str, Any], reviewed: bool) -> None:
    oid = opp.get("id")
    if not oid:
        return
    k = str(oid)

    cur = reg.get(k) or {}
    if not cur:
        cur["first_seen_iso"] = now_iso()
        cur["source"] = opp.get("source")
    cur["last_seen_iso"] = now_iso()
    cur["title"] = opp.get("title")
    cur["buyer"] = opp.get("buyer")
    cur["url"] = opp.get("url")
    cur["reviewed"] = bool(reviewed)

    reg[k] = cur

# -----------------------------
# Main ETL
# -----------------------------
def main() -> None:
    ensure_dirs()

    # env
    mp_ticket = os.getenv("MP_TICKET", "").strip()
    gh_token = os.getenv("GITHUB_TOKEN", "").strip()
    repo = os.getenv("GITHUB_REPOSITORY", "danielquinterosr/periscopio-compras-publicas").strip()

    candidates_top = safe_int(os.getenv("CANDIDATES_TOP", "800"), 800)
    max_detail = safe_int(os.getenv("MAX_DETAIL", "400"), 400)
    detail_sleep = float(os.getenv("DETAIL_SLEEP", "0.12"))

    # rules
    rules_all = load_rules("config/rules.yml") or {}
    defaults = rules_all.get("defaults") or {}
    by_source = rules_all.get("by_source") or {}

    # onu mapping
    onu_map = load_onu_mapping(DEFAULT_ONU_XLSX)
    deny_n1 = set(DEFAULT_DENY_NIVEL1)

    # reviewed ids
    reviewed_ids = fetch_reviewed_ids_from_github(repo=repo, token=gh_token) if gh_token else []
    reviewed_set = set(reviewed_ids)

    # registry
    registry = load_json_file(OUT_REGISTRY, {})
    if not isinstance(registry, dict):
        registry = {}

    # -----------------------------
    # 1) Cargar Compra Ágil desde XLSX
    # -----------------------------
    compra_xlsx = os.path.join(DATA_DIR, "compra_agil.xlsx")
    compra_rows = read_compra_agil_excel(compra_xlsx)

    # -----------------------------
    # 2) Cargar Licitaciones (listado)
    # -----------------------------
    lic_rows: List[Dict[str, Any]] = []
    lic_total = 0
    detalle_ok = 0
    detalle_fail = 0

    if mp_ticket:
        # paginación: recorremos hasta que venga vacío o hasta un límite defensivo
        page = 1
        empty_pages = 0
        while True:
            data = fetch_licitaciones_list(mp_ticket, page)
            listado = (data or {}).get("Listado") or []
            if not listado:
                empty_pages += 1
                if empty_pages >= 2:
                    break
                page += 1
                continue

            empty_pages = 0
            for it in listado:
                if not isinstance(it, dict):
                    continue

                codigo = (it.get("CodigoExterno") or it.get("Codigo") or "").strip()
                if not codigo:
                    continue

                lic_total += 1

                # normalización básica (puedes expandir según lo que ya tenías)
                lic_rows.append({
                    "source": "licitaciones",
                    "id": codigo,
                    "title": (it.get("Nombre") or it.get("NombreLicitacion") or "").strip(),
                    "buyer": (it.get("Comprador") or it.get("NombreOrganismo") or "").strip() or None,
                    "status": (it.get("Estado") or "").strip() or None,
                    "amount_clp": safe_float(it.get("MontoEstimado") or it.get("Monto") or it.get("MontoPresupuesto")),
                    "published_at": it.get("FechaPublicacion") or it.get("FechaCreacion") or None,
                    "questions_end_at": it.get("FechaCierrePreguntas") or None,
                    "close_at": it.get("FechaCierre") or None,
                    "url": f"https://www.mercadopublico.cl/Procurement/Modules/RFB/DetailsAcquisition.aspx?qs={codigo}",
                })

            # stop condition defensiva: evita loops eternos
            if page >= 200:
                break
            page += 1
    else:
        # sin ticket: no procesamos licitaciones
        lic_rows = []

    # -----------------------------
    # 3) Scoring + gating + show_min_score por fuente
    # -----------------------------
    all_rows = compra_rows + lic_rows

    scored: List[Dict[str, Any]] = []
    for o in all_rows:
        src = o.get("source") or "licitaciones"
        rules_src = (by_source.get(src) or {})
        # aplica defaults (thresholds) si faltan
        merged = {}
        merged.update(defaults)
        merged.update(rules_src)

        score, detail = total_score(
            text=f"{o.get('title','')} {o.get('buyer','')}",
            amount_clp=o.get("amount_clp"),
            rules=merged
        )
        o2 = dict(o)
        o2["score"] = int(score)
        o2["score_detail"] = detail
        scored.append(o2)

    # -----------------------------
    # 4) Detalle licitaciones (solo top candidates) para sacar CodigoCategoria
    #     y filtrar por Nivel1 ONU (anti-salud/equipos/eco/ultrasonido)
    # -----------------------------
    lic_scored = [x for x in scored if x.get("source") == "licitaciones"]
    lic_scored.sort(key=lambda x: x.get("score", 0), reverse=True)

    lic_candidates = lic_scored[:candidates_top]
    lic_kept: List[Dict[str, Any]] = []
    for idx, o in enumerate(lic_candidates):
        codigo = o.get("id")
        if not codigo:
            continue

        detail_json = fetch_licitacion_detail(mp_ticket, codigo) if mp_ticket else None
        if not detail_json:
            detalle_fail += 1
            # si no hay detalle, lo dejamos pasar (por ahora) sin categorías
            lic_kept.append(o)
            continue

        detalle_ok += 1
        cat_codes = extract_category_codes_from_detail(detail_json)
        if cat_codes:
            o["category_codes"] = cat_codes
            # map a nivel1
            nivel1s = []
            for c in cat_codes:
                info = onu_map.get(str(c))
                if info and info.get("Nivel1"):
                    nivel1s.append(info["Nivel1"])
            o["category_nivel1"] = sorted(set([x for x in nivel1s if x]))

            # filtro: si cualquier nivel1 cae en denylist, lo sacamos
            if set(o.get("category_nivel1", [])) & deny_n1:
                # descartado por rubro salud/lab
                pass
            else:
                lic_kept.append(o)
        else:
            # sin categorías: lo dejamos (y luego lo ajustamos cuando confirmemos estructura)
            lic_kept.append(o)

        if detail_sleep > 0:
            time.sleep(detail_sleep)

        if detalle_ok >= max_detail:
            # completamos el resto sin detalle
            lic_kept.extend(lic_candidates[idx+1:])
            break

    # reemplazamos licitaciones por la versión “kept”
    compra_scored = [x for x in scored if x.get("source") == "compra_agil"]
    scored = compra_scored + lic_kept

    # -----------------------------
    # 5) show_min_score + reviewed + registry
    # -----------------------------
    shown: List[Dict[str, Any]] = []
    for o in scored:
        src = o.get("source") or "licitaciones"
        rules_src = (by_source.get(src) or {})
        merged = {}
        merged.update(defaults)
        merged.update(rules_src)

        show_min = safe_int(((merged.get("thresholds") or {}).get("show_min_score", 2)), 2)
        if o.get("score", 0) < show_min:
            continue

        oid = str(o.get("id") or "")
        is_reviewed = oid in reviewed_set
        o["reviewed"] = bool(is_reviewed)

        upsert_registry(registry, o, reviewed=is_reviewed)
        shown.append(o)

    # sort final
    shown.sort(key=lambda x: (x.get("score", 0), x.get("published_at") or ""), reverse=True)

    # counts
    def count_src(xs: List[Dict[str, Any]], src: str) -> int:
        return sum(1 for x in xs if x.get("source") == src)

    def count_reviewed_src(xs: List[Dict[str, Any]], src: str) -> int:
        return sum(1 for x in xs if x.get("source") == src and bool(x.get("reviewed")))

    meta = {
        "last_update_iso": now_iso(),
        "repo": repo,
        "paths": {
            "compra_agil_xlsx": "data/compra_agil.xlsx",
            "registry": OUT_REGISTRY.replace(DOCS_DATA_DIR + "/", ""),
        },
        "counts": {
            "total_current": len(scored),
            "shown": len(shown),
            "licitaciones_total": lic_total,
            "compra_agil_total": len(compra_rows),
            "reviewed_ids_from_issues": len(reviewed_ids),
            "reviewed_licitaciones_shown": count_reviewed_src(shown, "licitaciones"),
            "reviewed_compra_agil_shown": count_reviewed_src(shown, "compra_agil"),
            "shown_licitaciones": count_src(shown, "licitaciones"),
            "shown_compra_agil": count_src(shown, "compra_agil"),
            "detalle_ok": detalle_ok,
            "detalle_fail": detalle_fail,
            "candidates_top": candidates_top,
            "max_detail": max_detail,
            "onu_rubros_loaded": bool(onu_map),
            "deny_nivel1": sorted(list(deny_n1)),
        },
        "version": "v0.8",
    }

    save_json_file(OUT_OPPS, shown)
    save_json_file(OUT_META, meta)
    save_json_file(OUT_REGISTRY, registry)

if __name__ == "__main__":
    main()
